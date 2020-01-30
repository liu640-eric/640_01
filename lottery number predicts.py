# -*- coding: utf-8 -*-
"""
Created on Fri Sep 20 10:08:48 2019

@author: liu640
"""

import operator # 把單維變多維 list
from functools import reduce # 把單維變多維 list 連結到第43列 
import openpyxl
book = openpyxl.load_workbook("D:\lottery_02.xlsx")
sheet = book.worksheets[0] #開啟第一個sheet
sheet2= book.worksheets[1] #開啟第二個sheet
sheet.insert_rows(2) #新增的號碼存在第二列，原第二列變第三列
no1st=1
lottery_no=[]
i=1
while i<8:
    a=int(input("樂透中獎號碼:"))
    lottery_no.append(a)
    i=i+1

################### 計算大樂透號碼單雙開出次數
j=0
k=0
l=0
while j<6:
    if (lottery_no[j]%2==0):
        k=k+1
    else:
        l=l+1
    j=j+1    
sheet.cell(2,10,l)
sheet.cell(2,11,k)

################### 計算大樂透號碼單雙開出次數

noth=int(input("樂透期數:"))
sheet.cell(2,1,noth)
for i in range(len(lottery_no)):
    sheet.cell(2,i+2,lottery_no[i])


lottery_matrix=[]
interval_total=noth-no1st+2 #用最新期數-第一期期數的區間值跑迴圈
for i in range(2,interval_total):
    for j in range(2, 8):
        lottery_matrix.append(sheet.cell(row=i, column=j).value)
" ".join(str(x) for x in lottery_matrix) #List 元素轉成文字

for i in range(1,50):
    stat=lottery_matrix.count(i)
    sheet2.cell(i+1,2,stat)


#以下為計算選擇一個號碼，則其他號碼出現的次數
#例如選擇9號，則出現9號的每一期的其他號碼出現的次數

def split_list (x): # 把單維變多維 list，然後6個號碼一組
   return [lottery_matrix[i:i+x] for i in range(0, len(lottery_matrix), x)]

lottery_matrix=split_list(6) #6個號碼一組

idno=input("標定預測號碼:")
lottery_set=[]  
for row in range(0,interval_total-2): #跑每一期的列數
    for col in range(0,6):#跑樂透6個號碼次數
        if (lottery_matrix[row][col]==int(idno)):
            lottery_set.append(lottery_matrix[row])
            
### 將多維矩陣改成單維 list
lottery_set_sep=reduce(operator.add, lottery_set) 

### 開始計算次數
for i in range(1,50):
    stat1=lottery_set_sep.count(i)
    sheet2.cell(i+1,3,stat1)
book.save(filename="D:\lottery_02.xlsx")